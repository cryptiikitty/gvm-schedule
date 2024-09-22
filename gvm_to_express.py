#!/usr/bin/python
# coding: utf-8
 
import sys
import base64
import hashlib
import hmac
import logging
import time
from datetime import datetime
from openpyxl import Workbook, load_workbook
 
import json
import requests
import gzip
import re
import datetime
import calendar
 
# Bot ID and URLs
bot_id = ''
bot_url = ''
secret = ""
sendto = ""
 
 
# Token
def get_token():
    token_url = '/api/v2/botx/bots/' + bot_id + '/token'
    h = hmac.new(secret.encode('utf-8'), bot_id.encode('utf-8'), hashlib.sha256)
    signature = base64.b16encode(h.digest())
    r = requests.get(bot_url + token_url, params={'signature': signature})
    return r.json()['result']
 
 
# API authorization
token = get_token()
if token:
    headers = {
        'authorization': 'Bearer ' + token,
        'content-type': 'application/json'
    }
else:
    logging.info(datetime.datetime.now(), 'Не получен токен', token)
 
# Message URL
msg_url = '/api/v3/botx/notification/callback/direct'
 
 
# Main function for sending messages to chat
def send_express(message):
    msg = message
    data = {
        "group_chat_id": sendto,
        "notification": {
            "status": "ok",
            "body": msg,
 
        }
    }
    try:
        requests.post(bot_url + msg_url, headers=headers, data=json.dumps(data))
    except Exception as ex:
        logging.info(repr(ex))
 
 
if __name__ == '__main__':
 
    excel_book = load_workbook('target.xlsx')
    sheets = excel_book.sheetnames
    ws = excel_book[sheets[0]]
 
    ex = (ws['A'][-1])
    ex = str(ex)
    start = ex.find("'.A") + len("'.A")
    end = ex.find('>')
    end_of_file = ex[start: end]
    end_of_file = int(end_of_file) + 1
    count = 1
 
    time_scan = ''
    ip_list = ''
 
    currentYear = datetime.datetime.now().year
    currentMonth = datetime.datetime.now().month
    currentDay = datetime.datetime.now().day
 
    date = datetime.datetime(currentYear, currentMonth, currentDay)
 
    j = 1
    while j < end_of_file:
        i = 1
 
        day_of_week_1 = calendar.day_name[date.weekday()]
        if day_of_week_1 == 'Friday':
            date += datetime.timedelta(days=3)
        else:
            date += datetime.timedelta(days=1)
 
        currentDate = str(date.year) + "-" + str(date.month) + "-" + str(date.day)
        isFriday = 0
        while i < end_of_file:
 
            day_of_week = calendar.day_name[date.weekday()]
            if day_of_week_1 == 'Friday':
                isFriday = 1
 
            cell = 'f' + str(i)
            my_values = ws[cell].value
            if str(my_values) == str(currentDate):
                cell_start_time = 'd' + str(i)
                start_time = ws[cell_start_time].value
                cell_end_time = 'e' + str(i)
                end_time = ws[cell_end_time].value
                time_scan = str(start_time) + ' - ' + str(end_time)
                cell_ip = 'c' + str(i)
                ip_list = ws[cell_ip].value
                scanDay = str(date.day) + "." + str(date.month) + "." + str(date.year)
                message = 'ℹ️Submitting nodes for scanning.' + '\n' + '-------------------------------------------------------' + '\n' + 'List of nodes: ' + ip_list + '\n' + 'Scan session time: ' + scanDay + ' ' + time_scan + '\n'
                send_express(message)
            i += 1
 
        if isFriday == 1:
            print('ждем 3 дня')
            time.sleep(259200)
        else:
            time.sleep(86400)
        j += 1
    excel_book.save('target.xlsx')
    excel_book.close()
