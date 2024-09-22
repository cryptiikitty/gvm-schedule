import sys
import os
from openpyxl import Workbook, load_workbook
import datetime
import time
import calendar
 
 
excel_book = load_workbook('target.xlsx')
sheets = excel_book.sheetnames
ws = excel_book[sheets[0]]
 
ex = (ws['A'][-1])
ex = str(ex)
end_of_file = ex[-2]
end_of_file = int(end_of_file) + 1
count = 1
 
#send to create scan task
 
j = 1
while j < end_of_file:
    cell = 'd' + str(j)
    schedule_time = ws[cell].value
    cell2 = 'f' + str(j)
    schedule_day = ws[cell2].value
    cell3 = 'b' + str(j)
    group_name = ws[cell3].value
    cell4 = 'c' + str(j)
    hosts = ws[cell4].value
    print(group_name)
    print('')
    command = "python3 pyGVM.py --hosts " + str(hosts) + " --time " + str(schedule_time) + " --day " + str(schedule_day) + " --name " + str(group_name)
    print('----------------------------')
    print('')
    time.sleep(5)
    os.system(command)
    j += 1
excel_book.save('target.xlsx')
excel_book.close()
