import sys
import os
from openpyxl import Workbook, load_workbook
import datetime
import time
import calendar
 
 
currentYear = datetime.datetime.now().year
currentMonth = datetime.datetime.now().month
currentDay = datetime.datetime.now().day
 
 
excel_book = load_workbook('target.xlsx')
sheets = excel_book.sheetnames
ws = excel_book[sheets[0]]
 
ex = (ws['A'][-1])
ex = str(ex)
 
start = ex.find("'.A") + len("'.A")
end = ex.find('>')
end_of_file = ex[start: end]
end_of_file = int(end_of_file) + 1
count = 1
 
 
date = datetime.datetime(currentYear,currentMonth,currentDay)
day_of_week = calendar.day_name[date.weekday()]
if day_of_week == 'Friday':
    date += datetime.timedelta(days=3)
else:
    date += datetime.timedelta(days=1)

 
while count < 100:
    i = 1
    while i < end_of_file:
 
        cell = 'a' + str(i)
        my_values = ws[cell].value
        if int(my_values) == count:
            cell2 = 'f' + str(i)
            date_start = str(date.year) + "-" + str(date.month) + "-" + str(date.day)
            ws[cell2].value = date_start
 
        i += 1
    day_of_week = calendar.day_name[date.weekday()]
    if day_of_week == 'Friday':
        date += datetime.timedelta(days=3)
    else:
        date += datetime.timedelta(days=1)
    count += 1
excel_book.save('target.xlsx')
excel_book.close()
